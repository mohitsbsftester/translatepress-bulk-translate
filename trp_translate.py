#!/usr/bin/env python3
"""
Bulk translation export, import and machine translation for TranslatePress.

TranslatePress stores page translations in `wp_trp_dictionary_<default>_<target>`,
one row per string:

    original      the source text, exactly as scraped from the rendered page
    translated    the translation
    status        0 untranslated, 1 machine, 2 human-reviewed

The table name encodes the language pair. With `default-language = ar` and a
target of `en_GB`, the active table is `wp_trp_dictionary_ar_en_gb` and a row
reads original = Arabic -> translated = English. Check yours under
Settings -> TranslatePress, and override with --default-lang / --target-lang.

The one thing to understand before using this: TranslatePress only renders a
translation when `original` matches, byte for byte, the string it scraped from
the page HTML. Hand-typed source text almost never matches, because the stored
copy carries HTML entities (&#8217;), non-breaking spaces, bidi marks and stray
diacritics. Hence the two-way workflow:

    1. export  - pull the real originals out of the DB into an .xlsx
    2. (human) - fill in the translation column
    3. import  - match the sheet back onto the DB and write translations

Round-tripping an export guarantees exact matches. A sheet written from scratch
still works, via the normalisation ladder in `norm_tiers`, but expect some rows
to land in the "unmatched" report.

Every command reads from either a live MySQL database or an offline .sql dump,
and every command defaults to a dry run.

    # Pull untranslated strings out for a translator
    ./trp_translate.py export --dump dump.sql --out todo.xlsx

    # See what a filled sheet would do, touching nothing
    ./trp_translate.py import --excel todo.xlsx --dump dump.sql

    # Machine-translate what is still empty, via OpenRouter
    ./trp_translate.py translate --dump dump.sql --sql-out patch.sql

    # Generate SQL to run through phpMyAdmin, plus its rollback
    ./trp_translate.py import --excel todo.xlsx --dump dump.sql \
        --sql-out patch.sql --backup rollback.sql

    # Write straight to a reachable database
    ./trp_translate.py import --excel todo.xlsx --wp-config wp-config.php --apply

Requires: openpyxl (always), pymysql (only for live database access).
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field

# TranslatePress status constants, from includes/queries/class-query.php
NOT_TRANSLATED = 0
MACHINE_TRANSLATED = 1
HUMAN_REVIEWED = 2
SIMILAR_TRANSLATED = 3

BLOCK_TYPE_REGULAR_STRING = 0

STATUS_NAMES = {
    NOT_TRANSLATED: "not translated",
    MACHINE_TRANSLATED: "machine translated",
    HUMAN_REVIEWED: "human reviewed",
    SIMILAR_TRANSLATED: "similar translated",
}


# --------------------------------------------------------------------------
# Text normalisation
# --------------------------------------------------------------------------

# Zero-width and bidirectional control characters. These are invisible, survive
# copy-paste into Excel, and are the single most common reason a hand-typed
# Arabic string fails to match the original scraped from the page.
INVISIBLES = re.compile(
    "["
    "​-‏"  # zero-width space/joiners, LRM, RLM
    "‪-‮"  # bidi embedding/override
    "⁠-⁤"  # word joiner, invisible operators
    "⁦-⁩"  # bidi isolates
    "﻿"  # BOM
    "]"
)

# Harakat (short vowels) and other combining marks. Optional in written Arabic,
# so the page and the spreadsheet frequently disagree about them.
ARABIC_DIACRITICS = re.compile(
    "["
    "ؐ-ؚ"
    "ً-ٟ"
    "ٰ"
    "ۖ-ۜ"
    "۟-ۨ"
    "۪-ۭ"
    "]"
)

TATWEEL = "ـ"  # kashida, a purely decorative letter-stretching character

# Letters with more than one common spelling.
ARABIC_LETTER_FOLD = str.maketrans(
    {
        "آ": "ا",  # آ -> ا
        "أ": "ا",  # أ -> ا
        "إ": "ا",  # إ -> ا
        "ٱ": "ا",  # ٱ -> ا
        "ى": "ي",  # ى -> ي
        "ة": "ه",  # ة -> ه
        "ی": "ي",  # Farsi yeh -> Arabic yeh
        "ک": "ك",  # Farsi keheh -> Arabic kaf
    }
)

# Arabic-Indic and Eastern Arabic-Indic digits -> ASCII.
DIGIT_FOLD = str.maketrans(
    {chr(0x0660 + i): str(i) for i in range(10)}
    | {chr(0x06F0 + i): str(i) for i in range(10)}
)

# Typographic punctuation that Excel/Word silently substitutes.
PUNCT_FOLD = str.maketrans(
    {
        "‘": "'",
        "’": "'",
        "‚": "'",
        "‛": "'",
        "“": '"',
        "”": '"',
        "„": '"',
        "–": "-",
        "—": "-",
        "−": "-",
        " ": " ",
        " ": " ",
        " ": " ",
        " ": " ",
    }
)

WHITESPACE = re.compile(r"\s+")


def unescape_all(text: str) -> str:
    """Fully decode HTML entities, including double-encoded ones."""
    for _ in range(3):
        decoded = html.unescape(text)
        if decoded == text:
            break
        text = decoded
    return text


def norm_tiers(text: str) -> list[str]:
    """
    Build progressively looser match keys for a string.

    Index the database at every tier, look the spreadsheet up at every tier, and
    take the tightest tier that hits. Returned keys are ordered strictest first:

        0  exact bytes
        1  Unicode NFC, invisibles stripped, trimmed
        2  + HTML entities decoded, punctuation folded, whitespace collapsed
        3  + Arabic diacritics/tatweel/letter-shape/digit folding, casefolded
    """
    if text is None:
        text = ""

    t0 = text

    t1 = unicodedata.normalize("NFC", text)
    t1 = INVISIBLES.sub("", t1)
    t1 = t1.strip()

    t2 = unescape_all(t1)
    t2 = t2.translate(PUNCT_FOLD)
    t2 = WHITESPACE.sub(" ", t2).strip()

    t3 = ARABIC_DIACRITICS.sub("", t2)
    t3 = t3.replace(TATWEEL, "")
    t3 = t3.translate(ARABIC_LETTER_FOLD)
    t3 = t3.translate(DIGIT_FOLD)
    t3 = t3.casefold()
    t3 = WHITESPACE.sub(" ", t3).strip()

    return [t0, t1, t2, t3]


TIER_NAMES = ["exact", "normalized", "entity-folded", "fuzzy"]

ARABIC_RANGE = re.compile(r"[؀-ۿݐ-ݿﭐ-﷿ﹰ-﻿]")
LATIN_RANGE = re.compile(r"[A-Za-z]")


def script_ratio(text: str, pattern: re.Pattern) -> float:
    """Fraction of non-space characters in `text` matching `pattern`."""
    meaningful = [c for c in text if not c.isspace() and not c.isdigit()]
    if not meaningful:
        return 0.0
    return sum(1 for c in meaningful if pattern.match(c)) / len(meaningful)


# --------------------------------------------------------------------------
# SQL helpers
# --------------------------------------------------------------------------


def sql_quote(value) -> str:
    """Quote a Python value as a MySQL literal."""
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    escaped = (
        str(value)
        .replace("\\", "\\\\")
        .replace("'", "\\'")
        .replace('"', '\\"')
        .replace("\n", "\\n")
        .replace("\r", "\\r")
        .replace("\x1a", "\\Z")
        .replace("\x00", "\\0")
    )
    return f"'{escaped}'"


def write_sql_file(path: str, statements: list[str], comments: list[str]) -> None:
    """
    Write a transaction-wrapped .sql patch fit for phpMyAdmin import.

    The `SET NAMES utf8mb4` line is not decoration: phpMyAdmin negotiates the
    connection charset from the file, and without it Arabic and typographic
    punctuation import as mojibake that looks like a translation bug days later.
    """
    with open(path, "w", encoding="utf-8") as handle:
        for line in comments:
            handle.write(f"-- {line}\n")
        handle.write(f"-- statements: {len(statements)}\n\n")
        handle.write("SET NAMES utf8mb4;\n")
        handle.write("SET SQL_MODE = 'NO_AUTO_VALUE_ON_ZERO';\n\n")
        handle.write("START TRANSACTION;\n\n")
        handle.write("\n".join(statements))
        handle.write("\n\nCOMMIT;\n")


def write_rollback(path: str, table: str, rows: list) -> None:
    """Emit a patch restoring every row's current translated text and status."""
    statements = [
        f"UPDATE `{table}` SET translated = {sql_quote(r.translated)}, "
        f"status = {r.status} WHERE id = {r.id};"
        for r in rows
    ]
    write_sql_file(path, statements, [f"Rollback snapshot of `{table}`"])


def sql_unescape(value: str) -> str:
    """Decode a MySQL string literal body back into a Python string."""
    mapping = {
        "n": "\n",
        "t": "\t",
        "r": "\r",
        "b": "\b",
        "0": "\0",
        "Z": "\x1a",
        "\\": "\\",
        "'": "'",
        '"': '"',
    }
    out: list[str] = []
    i = 0
    while i < len(value):
        char = value[i]
        if char == "\\" and i + 1 < len(value):
            nxt = value[i + 1]
            out.append(mapping.get(nxt, nxt))
            i += 2
        else:
            out.append(char)
            i += 1
    return "".join(out)


def split_sql_tuples(blob: str) -> list[list[str | None]]:
    """
    Split a MySQL `VALUES (...), (...)` blob into per-row field lists.

    Written by hand rather than with a regex because originals routinely contain
    commas, parentheses, escaped quotes and newlines. Returns raw field text;
    string fields are already unescaped, NULL becomes None.
    """
    rows: list[list[str | None]] = []
    current: list[str | None] = []
    buf: list[str] = []
    is_string = False
    in_string = False
    depth = 0
    i = 0

    def flush() -> None:
        raw = "".join(buf)
        if is_string:
            current.append(sql_unescape(raw))
        else:
            stripped = raw.strip()
            current.append(None if stripped.upper() == "NULL" else stripped)
        buf.clear()

    while i < len(blob):
        char = blob[i]

        if in_string:
            if char == "\\" and i + 1 < len(blob):
                buf.append(char)
                buf.append(blob[i + 1])
                i += 2
                continue
            if char == "'":
                # Doubled '' is a literal quote inside the string.
                if i + 1 < len(blob) and blob[i + 1] == "'":
                    buf.append("'")
                    i += 2
                    continue
                in_string = False
                i += 1
                continue
            buf.append(char)
            i += 1
            continue

        if char == "'":
            # Only inter-field whitespace can precede an opening quote; drop it
            # so the literal doesn't pick up the space after the comma.
            if not "".join(buf).strip():
                buf.clear()
            in_string = True
            is_string = True
            i += 1
            continue

        if char == "(" and depth == 0:
            depth = 1
            current = []
            buf.clear()
            is_string = False
            i += 1
            continue

        if depth == 1 and char == ",":
            flush()
            is_string = False
            i += 1
            continue

        if char == ")" and depth == 1:
            flush()
            rows.append(current)
            depth = 0
            is_string = False
            i += 1
            continue

        if depth == 1:
            buf.append(char)
        i += 1

    return rows


# --------------------------------------------------------------------------
# Data sources
# --------------------------------------------------------------------------


@dataclass
class DictRow:
    """One row of a wp_trp_dictionary_* table."""

    id: int
    original: str
    translated: str | None
    status: int
    block_type: int
    original_id: int | None


class Source:
    """Read-only view of the dictionary table."""

    def dictionary(self, table: str) -> list[DictRow]:
        raise NotImplementedError

    def close(self) -> None:
        pass


class DumpSource(Source):
    """Reads a mysqldump .sql file. Used for offline dry runs."""

    def __init__(self, path: str):
        self.path = path
        with open(path, encoding="utf-8", errors="replace") as handle:
            self.text = handle.read()

    def dictionary(self, table: str) -> list[DictRow]:
        pattern = re.compile(
            r"INSERT INTO `%s` \(([^)]*)\) VALUES\s*(.*?);\s*\n" % re.escape(table),
            re.S,
        )
        rows: list[DictRow] = []
        found_table = False

        for match in pattern.finditer(self.text):
            found_table = True
            columns = [c.strip().strip("`") for c in match.group(1).split(",")]
            index = {name: pos for pos, name in enumerate(columns)}
            for fields in split_sql_tuples(match.group(2)):
                if len(fields) != len(columns):
                    continue

                def get(name: str, default=None):
                    pos = index.get(name)
                    return fields[pos] if pos is not None else default

                rows.append(
                    DictRow(
                        id=int(get("id") or 0),
                        original=get("original") or "",
                        translated=get("translated"),
                        status=int(get("status") or 0),
                        block_type=int(get("block_type") or 0),
                        original_id=(
                            int(get("original_id")) if get("original_id") else None
                        ),
                    )
                )

        if not found_table and f"`{table}`" not in self.text:
            raise SystemExit(f"error: table `{table}` not found in {self.path}")
        return rows


class MySQLSource(Source):
    """Reads (and, via `execute`, writes) a live MySQL database."""

    def __init__(self, host: str, user: str, password: str, database: str, port: int):
        try:
            import pymysql
        except ImportError:
            raise SystemExit(
                "error: pymysql is required for live database access.\n"
                "       pip install pymysql"
            )
        self.conn = pymysql.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            port=port,
            charset="utf8mb4",
            autocommit=False,
        )

    def dictionary(self, table: str) -> list[DictRow]:
        with self.conn.cursor() as cur:
            cur.execute(
                f"SELECT id, original, translated, status, block_type, original_id "
                f"FROM `{table}`"
            )
            return [
                DictRow(
                    id=r[0],
                    original=r[1] or "",
                    translated=r[2],
                    status=r[3] or 0,
                    block_type=r[4] or 0,
                    original_id=r[5],
                )
                for r in cur.fetchall()
            ]

    def close(self) -> None:
        self.conn.close()


# --------------------------------------------------------------------------
# Machine translation via OpenRouter
# --------------------------------------------------------------------------

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
DEFAULT_MODEL = "google/gemini-3.1-flash-lite"

SYSTEM_PROMPT = """\
You are a professional Arabic-to-English translator working on website copy for \
{context}.

You receive a JSON object mapping string ids to Arabic source text. Return a JSON \
object with EXACTLY the same ids, mapping each to its English translation.

Rules:
- Output JSON only. No commentary, no markdown fences.
- Every input id must appear in the output. Never merge, split, drop or reorder.
- Preserve HTML tags and entities byte for byte: <strong>, <br />, &#8217;, &nbsp;, \
&amp;. Translate only the human-readable text between them.
- Preserve placeholders unchanged: %s, %1$s, {{name}}, [shortcode], {{{{tokens}}}}.
- Do not translate URLs, email addresses, phone numbers, or file names.
- Keep brand and proper names in their established English form.
- If a value is already English, return it unchanged.
- Match the register of marketing copy: natural, concise, not literal. Do not pad, \
do not add words that are not in the source.
- Preserve leading and trailing whitespace exactly as given.
"""


class TranslationError(RuntimeError):
    pass


class OpenRouterTranslator:
    """
    Batched Arabic -> English translation through OpenRouter.

    Strings are sent as an id-keyed JSON object rather than a list, because a
    model that drops or merges one line would silently shift every translation
    after it onto the wrong original. Ids make that failure detectable: any
    response missing an id is rejected and retried.
    """

    def __init__(self, api_key, model=DEFAULT_MODEL, context="a business website",
                 glossary=None, timeout=180, retries=3):
        self.api_key = api_key
        self.model = model
        self.context = context
        self.glossary = glossary or {}
        self.timeout = timeout
        self.retries = retries
        self.prompt_tokens = 0
        self.completion_tokens = 0

    def _system_prompt(self) -> str:
        prompt = SYSTEM_PROMPT.format(context=self.context)
        if self.glossary:
            terms = "\n".join(f"  {k} -> {v}" for k, v in self.glossary.items())
            prompt += (
                "\nUse these fixed translations wherever the term appears:\n" + terms + "\n"
            )
        return prompt

    def _post(self, payload: dict) -> dict:
        import urllib.error
        import urllib.request

        body = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            OPENROUTER_URL,
            data=body,
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
                "X-Title": "trp-translate",
            },
        )
        with urllib.request.urlopen(request, timeout=self.timeout) as response:
            return json.loads(response.read().decode("utf-8"))

    def translate_batch(self, items: list[tuple[str, str]]) -> dict[str, str]:
        """Translate [(id, arabic)] and return {id: english}."""
        import urllib.error

        source = {key: text for key, text in items}
        payload = {
            "model": self.model,
            "temperature": 0,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": self._system_prompt()},
                {"role": "user", "content": json.dumps(source, ensure_ascii=False)},
            ],
        }

        last_error = None
        for attempt in range(self.retries):
            try:
                data = self._post(payload)
                usage = data.get("usage") or {}
                self.prompt_tokens += usage.get("prompt_tokens", 0)
                self.completion_tokens += usage.get("completion_tokens", 0)

                content = data["choices"][0]["message"]["content"]
                # Some models wrap JSON in fences despite response_format.
                content = re.sub(r"^\s*```(?:json)?|```\s*$", "", content.strip())
                result = json.loads(content)
                if not isinstance(result, dict):
                    raise TranslationError("model returned a non-object")

                missing = [k for k in source if k not in result]
                if missing:
                    raise TranslationError(
                        f"{len(missing)} of {len(source)} ids missing from response"
                    )
                return {k: str(result[k]) for k in source}

            except (urllib.error.HTTPError, urllib.error.URLError, OSError) as exc:
                detail = ""
                if isinstance(exc, urllib.error.HTTPError):
                    try:
                        detail = ": " + exc.read().decode("utf-8", "replace")[:200]
                    except Exception:
                        pass
                    # Client errors other than rate limiting will not fix themselves.
                    if exc.code not in (408, 409, 429) and exc.code < 500:
                        raise TranslationError(f"HTTP {exc.code}{detail}") from exc
                last_error = TranslationError(f"{type(exc).__name__}{detail}")
            except (KeyError, ValueError, TranslationError) as exc:
                last_error = TranslationError(str(exc))

            if attempt < self.retries - 1:
                time.sleep(2 * (attempt + 1))

        raise last_error or TranslationError("translation failed")


def looks_translatable(text: str) -> bool:
    """
    True when a string is worth spending a model call on.

    Only strings actually containing Arabic qualify. Roughly 60% of this site's
    originals are leftover English from before the default language flipped to
    Arabic, and their correct English 'translation' is themselves - sending them
    to a model just invites it to rewrite copy nobody asked it to touch.
    """
    stripped = strip_tags(text).strip()
    if not stripped:
        return False
    if not ARABIC_RANGE.search(stripped):
        return False
    # Bare URLs, emails and numbers carry no translatable prose.
    if re.fullmatch(r"(https?://|www\.)\S+", stripped):
        return False
    if re.fullmatch(r"[^@\s]+@[^@\s]+\.[A-Za-z]{2,}", stripped):
        return False
    return True


def strip_tags(text: str) -> str:
    return re.sub(r"<[^>]+>", " ", text)


def read_wp_config(path: str) -> dict:
    """Pull DB credentials and $table_prefix out of a wp-config.php."""
    with open(path, encoding="utf-8", errors="replace") as handle:
        text = handle.read()

    config: dict = {}
    for key, const in (
        ("name", "DB_NAME"),
        ("user", "DB_USER"),
        ("password", "DB_PASSWORD"),
        ("host", "DB_HOST"),
    ):
        match = re.search(
            r"define\(\s*['\"]%s['\"]\s*,\s*['\"](.*?)['\"]\s*\)" % const, text
        )
        if match:
            config[key] = match.group(1)

    prefix = re.search(r"\$table_prefix\s*=\s*['\"](.*?)['\"]", text)
    config["prefix"] = prefix.group(1) if prefix else "wp_"

    host = config.get("host", "localhost")
    if ":" in host:
        host, _, port = host.partition(":")
        config["host"] = host
        if port.isdigit():
            config["port"] = int(port)

    return config


# --------------------------------------------------------------------------
# Spreadsheet I/O
# --------------------------------------------------------------------------


def require_openpyxl():
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "error: openpyxl is required to read and write .xlsx files.\n"
            "       pip install openpyxl"
        )
    return openpyxl


@dataclass
class SheetRow:
    number: int
    arabic: str
    english: str


def detect_columns(
    rows: list[list], arabic_hint: str | None, english_hint: str | None
) -> tuple[int, int, bool]:
    """
    Work out which column holds Arabic and which holds English.

    Returns (arabic_index, english_index, has_header). Explicit hints win; they
    accept a header name, a 1-based number, or a spreadsheet letter. Otherwise
    the header row is searched for the words "arabic"/"english", and failing
    that columns are classified by which script their content is actually in.
    """
    if not rows:
        raise SystemExit("error: the spreadsheet is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    width = max(len(r) for r in rows)

    def resolve(hint: str) -> int | None:
        if hint is None:
            return None
        hint = hint.strip()
        for pos, name in enumerate(header):
            if name.casefold() == hint.casefold():
                return pos
        if hint.isdigit():
            return int(hint) - 1
        if re.fullmatch(r"[A-Za-z]{1,2}", hint):
            value = 0
            for char in hint.upper():
                value = value * 26 + (ord(char) - 64)
            return value - 1
        raise SystemExit(f"error: cannot resolve column {hint!r}")

    arabic_index = resolve(arabic_hint)
    english_index = resolve(english_hint)

    header_looks_like_labels = False
    if arabic_index is None or english_index is None:
        for pos, name in enumerate(header):
            low = name.casefold()
            if arabic_index is None and re.search(r"arab|عرب", low):
                arabic_index = pos
                header_looks_like_labels = True
            if english_index is None and re.search(r"engl|انجل|إنجل", low):
                english_index = pos
                header_looks_like_labels = True

    if arabic_index is None or english_index is None:
        # Fall back to sniffing the content of each column.
        sample = rows[1:41] if header_looks_like_labels else rows[:40]
        scores = []
        for col in range(width):
            values = [
                str(r[col]) for r in sample if col < len(r) and r[col] not in (None, "")
            ]
            joined = " ".join(values)
            scores.append(
                (
                    script_ratio(joined, ARABIC_RANGE),
                    script_ratio(joined, LATIN_RANGE),
                    len(values),
                )
            )
        if arabic_index is None:
            candidates = [
                (s[0], i) for i, s in enumerate(scores) if s[2] and i != english_index
            ]
            arabic_index = max(candidates)[1] if candidates else None
        if english_index is None:
            candidates = [
                (s[1], i) for i, s in enumerate(scores) if s[2] and i != arabic_index
            ]
            english_index = max(candidates)[1] if candidates else None

    if arabic_index is None or english_index is None or arabic_index == english_index:
        raise SystemExit(
            "error: could not identify the Arabic and English columns.\n"
            "       Pass --arabic-col and --english-col explicitly "
            "(header name, number, or letter)."
        )

    # Treat row 1 as a header if it is non-empty and clearly isn't data.
    has_header = header_looks_like_labels
    if not has_header and len(rows) > 1:
        first = rows[0]
        first_arabic = (
            str(first[arabic_index])
            if arabic_index < len(first) and first[arabic_index]
            else ""
        )
        has_header = bool(first_arabic) and script_ratio(first_arabic, ARABIC_RANGE) < 0.2

    return arabic_index, english_index, has_header


def load_sheet(
    path: str, sheet: str | None, arabic_hint: str | None, english_hint: str | None
) -> list[SheetRow]:
    """Read the Arabic/English pairs out of a spreadsheet."""
    if path.lower().endswith((".csv", ".tsv")):
        delimiter = "\t" if path.lower().endswith(".tsv") else ","
        with open(path, encoding="utf-8-sig", newline="") as handle:
            raw = [row for row in csv.reader(handle, delimiter=delimiter)]
    else:
        openpyxl = require_openpyxl()
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet:
            if sheet not in book.sheetnames:
                raise SystemExit(
                    f"error: no sheet named {sheet!r}. "
                    f"Available: {', '.join(book.sheetnames)}"
                )
            worksheet = book[sheet]
        else:
            worksheet = book[book.sheetnames[0]]
        raw = [list(r) for r in worksheet.iter_rows(values_only=True)]
        book.close()

    raw = [r for r in raw if any(c not in (None, "") for c in r)]
    if not raw:
        raise SystemExit("error: the spreadsheet has no data rows")

    arabic_index, english_index, has_header = detect_columns(
        raw, arabic_hint, english_hint
    )

    def cell(row: list, index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        return str(row[index])

    result: list[SheetRow] = []
    for offset, row in enumerate(raw):
        if offset == 0 and has_header:
            continue
        arabic = cell(row, arabic_index)
        english = cell(row, english_index)
        if not arabic.strip() and not english.strip():
            continue
        result.append(SheetRow(number=offset + 1, arabic=arabic, english=english))

    sys.stderr.write(
        f"  columns: Arabic=#{arabic_index + 1} English=#{english_index + 1} "
        f"header_row={'yes' if has_header else 'no'}\n"
    )
    return result


# --------------------------------------------------------------------------
# Matching
# --------------------------------------------------------------------------


@dataclass
class Match:
    row: SheetRow
    tier: int | None = None
    target: DictRow | None = None
    outcome: str = "unmatched"
    note: str = ""


def build_index(rows: list[DictRow]) -> list[dict[str, list[DictRow]]]:
    """Index dictionary rows by every normalisation tier."""
    index: list[dict[str, list[DictRow]]] = [defaultdict(list) for _ in TIER_NAMES]
    for row in rows:
        for tier, key in enumerate(norm_tiers(row.original)):
            index[tier][key].append(row)
    return index


def match_rows(
    sheet: list[SheetRow],
    dictionary: list[DictRow],
    max_tier: int,
    skip_translated: bool,
    status: int,
) -> list[Match]:
    """Match each spreadsheet row onto a dictionary row, tightest tier first."""
    index = build_index(dictionary)
    results: list[Match] = []

    for row in sheet:
        match = Match(row=row)
        english = row.english.strip()
        keys = norm_tiers(row.arabic)

        for tier in range(min(max_tier + 1, len(TIER_NAMES))):
            candidates = index[tier].get(keys[tier])
            if not candidates:
                continue

            distinct = {c.id: c for c in candidates}
            if len(distinct) > 1:
                # Same normalised text, several dictionary rows. Prefer an
                # untranslated one so we fill a gap rather than churn a
                # translation someone already reviewed.
                empty = [c for c in distinct.values() if not (c.translated or "").strip()]
                chosen = (empty or list(distinct.values()))[0]
                match.note = (
                    f"{len(distinct)} dictionary rows share this original "
                    f"(ids {', '.join(str(i) for i in sorted(distinct))}); "
                    f"chose id {chosen.id}"
                )
                match.outcome = "ambiguous"
            else:
                chosen = next(iter(distinct.values()))

            match.tier = tier
            match.target = chosen
            break

        if match.target is None:
            match.outcome = "unmatched"
            results.append(match)
            continue

        if not english:
            match.outcome = "skipped"
            match.note = "no English text in the spreadsheet"
        elif skip_translated and (match.target.translated or "").strip():
            match.outcome = "skipped"
            match.note = "already translated (--skip-translated)"
        elif (match.target.translated or "") == english:
            if match.target.status == status:
                match.outcome = "unchanged"
            else:
                # Same text, different status - e.g. promoting a machine
                # translation to human-reviewed. Worth doing, but it is not an
                # overwrite and shouldn't be reported as one.
                match.outcome = "status-only"
                match.note = (
                    f"text unchanged; status "
                    f"{STATUS_NAMES.get(match.target.status, match.target.status)}"
                    f" -> {STATUS_NAMES.get(status, status)}"
                )
        elif (match.target.translated or "").strip():
            if match.outcome != "ambiguous":
                match.outcome = "overwrite"
            match.note = (
                match.note + "; " if match.note else ""
            ) + f"replaces {(match.target.translated or '')[:60]!r}"
        else:
            if match.outcome != "ambiguous":
                match.outcome = "fill"

        results.append(match)

    return results


# --------------------------------------------------------------------------
# Commands
# --------------------------------------------------------------------------


def open_source(args) -> tuple[Source, str]:
    """Open the configured source and return it alongside the table prefix."""
    prefix = args.prefix

    if args.dump:
        return DumpSource(args.dump), prefix

    config: dict = {}
    if args.wp_config:
        config = read_wp_config(args.wp_config)
        if not args.prefix_explicit:
            prefix = config.get("prefix", prefix)

    host = args.db_host or config.get("host") or "localhost"
    user = args.db_user or config.get("user")
    password = args.db_pass if args.db_pass is not None else config.get("password")
    database = args.db_name or config.get("name")
    port = args.db_port or config.get("port") or 3306

    if not (user and database):
        raise SystemExit(
            "error: no source configured.\n"
            "       Use --dump FILE.sql for offline work, or supply database\n"
            "       credentials via --wp-config or --db-name/--db-user/--db-pass."
        )

    return MySQLSource(host, user, password, database, port), prefix


def table_name(prefix: str, default_lang: str, target_lang: str) -> str:
    return f"{prefix}trp_dictionary_{default_lang.lower()}_{target_lang.lower()}"


def cmd_export(args) -> int:
    source, prefix = open_source(args)
    table = args.table or table_name(prefix, args.default_lang, args.target_lang)
    sys.stderr.write(f"Reading `{table}`\n")

    rows = source.dictionary(table)
    source.close()

    if not args.all:
        rows = [r for r in rows if not (r.translated or "").strip()]
    rows.sort(key=lambda r: r.id)

    if not rows:
        sys.stderr.write("Nothing to export.\n")
        return 0

    if args.out.lower().endswith((".csv", ".tsv")):
        delimiter = "\t" if args.out.lower().endswith(".tsv") else ","
        with open(args.out, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerow(["id", "Arabic (original)", "English (translation)", "status"])
            for row in rows:
                writer.writerow(
                    [
                        row.id,
                        row.original,
                        row.translated or "",
                        STATUS_NAMES.get(row.status, row.status),
                    ]
                )
    else:
        openpyxl = require_openpyxl()
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "translations"
        sheet.append(["id", "Arabic (original)", "English (translation)", "status"])

        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="333333")
        for cell in sheet[1]:
            cell.font = head_font
            cell.fill = head_fill

        for row in rows:
            sheet.append(
                [
                    row.id,
                    row.original,
                    row.translated or "",
                    STATUS_NAMES.get(row.status, row.status),
                ]
            )

        for column, width in ((1, 8), (2, 70), (3, 70), (4, 18)):
            sheet.column_dimensions[get_column_letter(column)].width = width

        # Arabic reads right-to-left; English does not.
        for record in sheet.iter_rows(min_row=2):
            record[1].alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="right", readingOrder=2
            )
            record[2].alignment = Alignment(wrap_text=True, vertical="top")

        sheet.freeze_panes = "A2"
        book.save(args.out)

    sys.stderr.write(f"Wrote {len(rows)} row(s) to {args.out}\n")
    sys.stderr.write(
        "Fill in the English column and re-import. Leave the Arabic column\n"
        "untouched so every row matches exactly on the way back in.\n"
    )
    return 0


def cmd_import(args) -> int:
    source, prefix = open_source(args)
    table = args.table or table_name(prefix, args.default_lang, args.target_lang)

    sys.stderr.write(f"Reading {args.excel}\n")
    sheet = load_sheet(args.excel, args.sheet, args.arabic_col, args.english_col)
    sys.stderr.write(f"  {len(sheet)} data row(s)\n\n")

    sys.stderr.write(f"Reading `{table}`\n")
    dictionary = source.dictionary(table)
    translated_before = sum(1 for r in dictionary if (r.translated or "").strip())
    sys.stderr.write(
        f"  {len(dictionary)} row(s), {translated_before} translated, "
        f"{len(dictionary) - translated_before} empty\n\n"
    )

    matches = match_rows(
        sheet, dictionary, args.max_tier, args.skip_translated, args.status
    )

    writes = [
        m
        for m in matches
        if m.outcome in ("fill", "overwrite", "ambiguous", "status-only")
    ]
    unmatched = [m for m in matches if m.outcome == "unmatched"]

    # Two spreadsheet rows can normalise onto the same dictionary row. Emitting
    # both UPDATEs would make the result depend on statement order, so keep the
    # first and report the rest rather than silently letting the last one win.
    seen: dict[int, Match] = {}
    conflicts: list[tuple[Match, Match]] = []
    deduped: list[Match] = []
    for match in writes:
        previous = seen.get(match.target.id)
        if previous is None:
            seen[match.target.id] = match
            deduped.append(match)
        elif previous.row.english.strip() != match.row.english.strip():
            conflicts.append((previous, match))
            match.outcome = "conflict"
            match.note = (
                f"row {previous.row.number} already maps to dictionary id "
                f"{match.target.id} with a different translation; skipped"
            )
    writes = deduped

    # Counted after dedup so the summary reflects what will actually be written.
    counts: dict[str, int] = defaultdict(int)
    tiers: dict[int, int] = defaultdict(int)
    for match in matches:
        counts[match.outcome] += 1
        if match.tier is not None:
            tiers[match.tier] += 1

    sys.stderr.write("Match results\n")
    for outcome in (
        "fill",
        "overwrite",
        "status-only",
        "unchanged",
        "ambiguous",
        "conflict",
        "skipped",
        "unmatched",
    ):
        if counts[outcome]:
            sys.stderr.write(f"  {outcome:<12} {counts[outcome]:>5}\n")
    if tiers:
        detail = "  ".join(f"{TIER_NAMES[t]}={tiers[t]}" for t in sorted(tiers))
        sys.stderr.write(f"  matched via   {detail}\n")
    sys.stderr.write("\n")

    if conflicts:
        sys.stderr.write(
            f"{len(conflicts)} row(s) collide with an earlier row on the same "
            "dictionary\nentry but disagree on the English. The earlier row wins; "
            "the rest are skipped:\n"
        )
        for previous, match in conflicts[:10]:
            sys.stderr.write(
                f"  id {match.target.id}: row {previous.row.number} "
                f"{previous.row.english[:34]!r} vs row {match.row.number} "
                f"{match.row.english[:34]!r}\n"
            )
        if len(conflicts) > 10:
            sys.stderr.write(f"  ... and {len(conflicts) - 10} more\n")
        sys.stderr.write("\n")

    if unmatched:
        sys.stderr.write(
            f"{len(unmatched)} row(s) matched no original in `{table}`. "
            "TranslatePress\n"
            "only substitutes text it has already scraped from a page, so these\n"
            "will have no effect until the exact source string exists:\n"
        )
        for match in unmatched[:10]:
            sys.stderr.write(f"  row {match.row.number}: {match.row.arabic[:70]!r}\n")
        if len(unmatched) > 10:
            sys.stderr.write(f"  ... and {len(unmatched) - 10} more\n")
        sys.stderr.write(
            "Run `export` to pull the real originals out and fill those in instead.\n\n"
        )

    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "sheet_row",
                    "outcome",
                    "matched_via",
                    "dictionary_id",
                    "arabic",
                    "english",
                    "previous_translation",
                    "note",
                ]
            )
            for match in matches:
                writer.writerow(
                    [
                        match.row.number,
                        match.outcome,
                        TIER_NAMES[match.tier] if match.tier is not None else "",
                        match.target.id if match.target else "",
                        match.row.arabic,
                        match.row.english,
                        (match.target.translated or "") if match.target else "",
                        match.note,
                    ]
                )
        sys.stderr.write(f"Full report written to {args.report}\n\n")

    if not writes:
        sys.stderr.write("No changes to write.\n")
        source.close()
        return 0

    statements = [
        (
            f"UPDATE `{table}` SET translated = {sql_quote(m.row.english.strip())}, "
            f"status = {args.status} WHERE id = {m.target.id};"
        )
        for m in writes
    ]

    # Written before any branch: with no shell on the server, --sql-out is the
    # only way to apply changes, and that path still deserves a rollback.
    if args.backup:
        write_rollback(args.backup, table, dictionary)
        sys.stderr.write(f"Rollback script written to {args.backup}\n")

    if args.sql_out:
        write_sql_file(
            args.sql_out,
            statements,
            [
                "TranslatePress bulk translation import",
                f"table: {table}",
                f"source: {args.excel}",
            ],
        )
        sys.stderr.write(f"Wrote {len(statements)} UPDATE(s) to {args.sql_out}\n")
        source.close()
        return 0

    if not args.apply:
        sys.stderr.write(
            f"DRY RUN - {len(statements)} row(s) would be updated. "
            "Nothing was written.\n"
        )
        for match in writes[:5]:
            sys.stderr.write(
                f"  id {match.target.id}: {match.row.arabic[:40]!r} "
                f"-> {match.row.english[:40]!r}\n"
            )
        if len(writes) > 5:
            sys.stderr.write(f"  ... and {len(writes) - 5} more\n")
        sys.stderr.write("\nRe-run with --apply to write, or --sql-out FILE for SQL.\n")
        source.close()
        return 0

    if not isinstance(source, MySQLSource):
        raise SystemExit(
            "error: --apply needs a live database connection, not --dump.\n"
            "       Use --sql-out to generate SQL from a dump instead."
        )

    with source.conn.cursor() as cur:
        for match in writes:
            cur.execute(
                f"UPDATE `{table}` SET translated = %s, status = %s WHERE id = %s",
                (match.row.english.strip(), args.status, match.target.id),
            )
    source.conn.commit()
    sys.stderr.write(f"Applied {len(writes)} update(s) to `{table}`.\n")
    sys.stderr.write(
        "Clear any page/object cache and CDN so the new strings are served.\n"
    )
    source.close()
    return 0


def cmd_translate(args) -> int:
    source, prefix = open_source(args)
    table = args.table or table_name(prefix, args.default_lang, args.target_lang)

    api_key = args.api_key or os.environ.get("OPENROUTER_API_KEY")
    if not api_key and not args.estimate_only:
        raise SystemExit(
            "error: no OpenRouter API key.\n"
            "       export OPENROUTER_API_KEY=sk-or-... or pass --api-key\n"
            "       (use --estimate-only to price the job without a key)"
        )

    glossary = {}
    if args.glossary:
        with open(args.glossary, encoding="utf-8") as handle:
            glossary = json.load(handle)
        sys.stderr.write(f"Glossary: {len(glossary)} fixed term(s)\n")

    sys.stderr.write(f"Reading `{table}`\n")
    dictionary = source.dictionary(table)
    sys.stderr.write(f"  {len(dictionary)} row(s)\n")

    if args.retranslate:
        pool = list(dictionary)
    else:
        pool = [r for r in dictionary if not (r.translated or "").strip()]

    candidates = [r for r in pool if looks_translatable(r.original)]
    skipped = len(pool) - len(candidates)
    sys.stderr.write(
        f"  {len(pool)} candidate(s), {len(candidates)} contain Arabic, "
        f"{skipped} skipped as non-Arabic\n"
    )
    if args.limit:
        candidates = candidates[: args.limit]
        sys.stderr.write(f"  limited to first {len(candidates)}\n")

    if not candidates:
        sys.stderr.write("\nNothing to translate.\n")
        source.close()
        return 0

    chars = sum(len(r.original) for r in candidates)
    # Arabic runs ~2.5 chars/token; add the system prompt once per batch and
    # assume output is roughly the size of input.
    batches = (len(candidates) + args.batch_size - 1) // args.batch_size
    est_in = chars / 2.5 + batches * 320
    est_out = chars / 2.5
    est_cost = (est_in * args.price_in + est_out * args.price_out) / 1e6
    sys.stderr.write(
        f"\n{len(candidates)} string(s), {chars:,} chars, {batches} batch(es)\n"
        f"Model: {args.model}\n"
        f"Estimated: ~{est_in/1000:.1f}k in + ~{est_out/1000:.1f}k out tokens "
        f"= ~${est_cost:.4f}\n\n"
    )

    if args.estimate_only:
        source.close()
        return 0
    if est_cost > args.max_cost:
        raise SystemExit(
            f"error: estimated ${est_cost:.4f} exceeds --max-cost ${args.max_cost:.2f}.\n"
            f"       Raise the cap or use --limit to translate a subset."
        )

    translator = OpenRouterTranslator(
        api_key, args.model, args.context, glossary, args.timeout
    )

    results: dict[int, str] = {}
    failures: list[tuple[DictRow, str]] = []
    for index in range(0, len(candidates), args.batch_size):
        chunk = candidates[index : index + args.batch_size]
        items = [(str(r.id), r.original) for r in chunk]
        label = f"batch {index // args.batch_size + 1}/{batches}"
        try:
            got = translator.translate_batch(items)
            for row in chunk:
                value = got.get(str(row.id), "").strip()
                if value:
                    results[row.id] = value
            sys.stderr.write(f"  {label}: {len(chunk)} ok\n")
        except TranslationError as exc:
            # One bad batch shouldn't sink the run; retry its strings alone so a
            # single unparseable string can't cost the other 24 in the chunk.
            sys.stderr.write(f"  {label}: FAILED ({exc}) - retrying individually\n")
            for row in chunk:
                try:
                    got = translator.translate_batch([(str(row.id), row.original)])
                    value = got.get(str(row.id), "").strip()
                    if value:
                        results[row.id] = value
                except TranslationError as inner:
                    failures.append((row, str(inner)))

    spent = (
        translator.prompt_tokens * args.price_in
        + translator.completion_tokens * args.price_out
    ) / 1e6
    sys.stderr.write(
        f"\nTranslated {len(results)}/{len(candidates)}. "
        f"Tokens: {translator.prompt_tokens:,} in / "
        f"{translator.completion_tokens:,} out. Actual cost ~${spent:.4f}\n"
    )
    if failures:
        sys.stderr.write(f"{len(failures)} string(s) failed:\n")
        for row, why in failures[:5]:
            sys.stderr.write(f"  id {row.id}: {why} - {row.original[:50]!r}\n")
        if len(failures) > 5:
            sys.stderr.write(f"  ... and {len(failures) - 5} more\n")

    by_id = {r.id: r for r in candidates}
    if args.report:
        with open(args.report, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["id", "arabic", "english", "previous"])
            for rid, english in sorted(results.items()):
                writer.writerow(
                    [rid, by_id[rid].original, english, by_id[rid].translated or ""]
                )
        sys.stderr.write(f"Report written to {args.report}\n")

    if not results:
        source.close()
        return 1

    statements = [
        f"UPDATE `{table}` SET translated = {sql_quote(text)}, "
        f"status = {args.status} WHERE id = {rid};"
        for rid, text in sorted(results.items())
    ]

    if args.backup:
        write_rollback(args.backup, table, dictionary)
        sys.stderr.write(f"Rollback script written to {args.backup}\n")

    if args.sql_out:
        write_sql_file(
            args.sql_out,
            statements,
            [f"Machine translation via {args.model}", f"table: {table}"],
        )
        sys.stderr.write(f"Wrote {len(statements)} UPDATE(s) to {args.sql_out}\n")
        source.close()
        return 0

    if not args.apply:
        sys.stderr.write(f"\nDRY RUN - {len(results)} row(s) would be updated.\n")
        for rid, text in list(sorted(results.items()))[:5]:
            sys.stderr.write(
                f"  id {rid}: {by_id[rid].original[:40]!r} -> {text[:40]!r}\n"
            )
        if len(results) > 5:
            sys.stderr.write(f"  ... and {len(results) - 5} more\n")
        sys.stderr.write("\nRe-run with --apply to write, or --sql-out FILE for SQL.\n")
        source.close()
        return 0

    if not isinstance(source, MySQLSource):
        raise SystemExit(
            "error: --apply needs a live database connection, not --dump.\n"
            "       Use --sql-out to generate SQL from a dump instead."
        )

    with source.conn.cursor() as cur:
        for rid, text in sorted(results.items()):
            cur.execute(
                f"UPDATE `{table}` SET translated = %s, status = %s WHERE id = %s",
                (text, args.status, rid),
            )
    source.conn.commit()
    sys.stderr.write(f"Applied {len(results)} translation(s) to `{table}`.\n")
    sys.stderr.write(
        "These are marked machine-translated - review them in the "
        "TranslatePress editor.\n"
    )
    source.close()
    return 0


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------


def add_source_args(parser: argparse.ArgumentParser) -> None:
    group = parser.add_argument_group("source")
    group.add_argument("--dump", help="read from a mysqldump .sql file (offline)")
    group.add_argument("--wp-config", help="read DB credentials from a wp-config.php")
    group.add_argument("--db-host", help="MySQL host")
    group.add_argument("--db-port", type=int, help="MySQL port (default 3306)")
    group.add_argument("--db-name", help="MySQL database name")
    group.add_argument("--db-user", help="MySQL user")
    group.add_argument("--db-pass", help="MySQL password")

    group = parser.add_argument_group("table")
    group.add_argument("--prefix", default="wp_", help="table prefix (default: wp_)")
    group.add_argument(
        "--default-lang",
        default="ar",
        help="TranslatePress default language (default: ar)",
    )
    group.add_argument(
        "--target-lang",
        default="en_GB",
        help="translation language to write (default: en_GB)",
    )
    group.add_argument("--table", help="override the dictionary table name entirely")


def use_utf8_console() -> None:
    """
    Force UTF-8 on the console streams.

    Progress output echoes source strings, so on Windows - where the console
    still defaults to a legacy code page - printing Arabic raises
    UnicodeEncodeError and kills the run partway through. errors='replace'
    means an unrepresentable glyph degrades to '?' instead of aborting.
    """
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except (ValueError, OSError):
                pass


def main(argv: list[str]) -> int:
    use_utf8_console()
    parser = argparse.ArgumentParser(
        prog="trp_translate.py",
        description="Bulk import and export TranslatePress translations.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__.split("Requires:")[0].split("Both commands")[1].strip()
        if "Both commands" in __doc__
        else None,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    exporter = sub.add_parser(
        "export", help="write untranslated originals to a spreadsheet"
    )
    add_source_args(exporter)
    exporter.add_argument(
        "--out", default="translations.xlsx", help="output .xlsx or .csv"
    )
    exporter.add_argument(
        "--all",
        action="store_true",
        help="include rows that already have a translation",
    )
    exporter.set_defaults(func=cmd_export)

    importer = sub.add_parser(
        "import", help="read a spreadsheet and write translations back"
    )
    add_source_args(importer)
    importer.add_argument("--excel", required=True, help="input .xlsx, .csv or .tsv")
    importer.add_argument("--sheet", help="worksheet name (default: first)")
    importer.add_argument(
        "--arabic-col", help="Arabic column: header name, number, or letter"
    )
    importer.add_argument(
        "--english-col", help="English column: header name, number, or letter"
    )
    importer.add_argument(
        "--status",
        type=int,
        default=HUMAN_REVIEWED,
        choices=[NOT_TRANSLATED, MACHINE_TRANSLATED, HUMAN_REVIEWED, SIMILAR_TRANSLATED],
        help="status to set (default: 2, human reviewed)",
    )
    importer.add_argument(
        "--max-tier",
        type=int,
        default=2,
        choices=[0, 1, 2, 3],
        help=(
            "loosest matching tier: 0 exact, 1 normalized, 2 entity-folded "
            "(default), 3 fuzzy Arabic folding"
        ),
    )
    importer.add_argument(
        "--skip-translated",
        action="store_true",
        help="leave rows that already have a translation alone",
    )
    importer.add_argument("--report", help="write a per-row CSV report here")
    importer.add_argument("--sql-out", help="write UPDATE statements here instead of applying")
    importer.add_argument(
        "--backup", help="write a rollback .sql of current values before applying"
    )
    importer.add_argument(
        "--apply", action="store_true", help="actually write to the database"
    )
    importer.set_defaults(func=cmd_import)

    translator = sub.add_parser(
        "translate", help="machine-translate untranslated strings via OpenRouter"
    )
    add_source_args(translator)
    translator.add_argument(
        "--api-key", help="OpenRouter key (default: $OPENROUTER_API_KEY)"
    )
    translator.add_argument(
        "--model", default=DEFAULT_MODEL, help=f"model id (default: {DEFAULT_MODEL})"
    )
    translator.add_argument(
        "--context",
        default="a business website",
        help=(
            "one-line description of the site, used to steer tone. Worth setting: "
            "'a commercial printing press in Riyadh' yields better copy than the default"
        ),
    )
    translator.add_argument(
        "--glossary", help="JSON file of {arabic: english} terms to pin"
    )
    translator.add_argument(
        "--batch-size", type=int, default=25, help="strings per request (default: 25)"
    )
    translator.add_argument(
        "--limit", type=int, help="translate at most N strings (use for a test run)"
    )
    translator.add_argument(
        "--retranslate",
        action="store_true",
        help="include rows that already have a translation",
    )
    translator.add_argument(
        "--status",
        type=int,
        default=MACHINE_TRANSLATED,
        choices=[NOT_TRANSLATED, MACHINE_TRANSLATED, HUMAN_REVIEWED, SIMILAR_TRANSLATED],
        help="status to set (default: 1, machine translated)",
    )
    translator.add_argument(
        "--price-in", type=float, default=0.125, help="$ per 1M input tokens"
    )
    translator.add_argument(
        "--price-out", type=float, default=0.75, help="$ per 1M output tokens"
    )
    translator.add_argument(
        "--max-cost", type=float, default=5.0, help="abort if estimate exceeds this"
    )
    translator.add_argument(
        "--estimate-only", action="store_true", help="price the job and exit"
    )
    translator.add_argument("--timeout", type=int, default=180, help="HTTP timeout (s)")
    translator.add_argument("--report", help="write a per-string CSV report here")
    translator.add_argument("--sql-out", help="write UPDATE statements here")
    translator.add_argument("--backup", help="write a rollback .sql before applying")
    translator.add_argument(
        "--apply", action="store_true", help="actually write to the database"
    )
    translator.set_defaults(func=cmd_translate)

    args = parser.parse_args(argv)
    args.prefix_explicit = "--prefix" in argv
    return args.func(args)


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv[1:]))
    except KeyboardInterrupt:
        sys.exit(130)
