"""Human-review CSV and XLSX report output."""

from __future__ import annotations

import csv
from pathlib import Path

from .models import REPORT_FIELDS, TranslationRecord
from .spreadsheet import require_openpyxl


def write_review(path: str | Path, records: list[TranslationRecord]) -> None:
    destination = Path(path)
    rows = [record.as_report_row() for record in records]
    if destination.suffix.casefold() in (".csv", ".tsv"):
        delimiter = "\t" if destination.suffix.casefold() == ".tsv" else ","
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=REPORT_FIELDS, delimiter=delimiter
            )
            writer.writeheader()
            writer.writerows(rows)
        return
    if destination.suffix.casefold() != ".xlsx":
        raise ValueError("review report must end in .csv, .tsv, or .xlsx")

    openpyxl = require_openpyxl()
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(REPORT_FIELDS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
    for row in rows:
        sheet.append([row[field] for field in REPORT_FIELDS])
    widths = {
        "A": 10,
        "B": 18,
        "C": 18,
        "D": 65,
        "E": 65,
        "F": 35,
        "G": 20,
        "H": 20,
        "I": 20,
        "J": 18,
        "K": 22,
        "L": 20,
        "M": 45,
        "N": 55,
        "O": 68,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for record in sheet.iter_rows(min_row=2):
        for cell in record:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row > 1:
        failed_fill = PatternFill("solid", fgColor="FCE8E6")
        warning_fill = PatternFill("solid", fgColor="FFF4CC")
        sheet.conditional_formatting.add(
            f"A2:O{sheet.max_row}",
            FormulaRule(formula=['$L2="failed"'], fill=failed_fill),
        )
        sheet.conditional_formatting.add(
            f"A2:O{sheet.max_row}",
            FormulaRule(formula=['$M2<>""'], fill=warning_fill),
        )
    workbook.save(destination)
