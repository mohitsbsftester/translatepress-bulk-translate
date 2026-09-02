"""Language-neutral CSV/XLSX export, import, and dictionary matching."""

from __future__ import annotations

import csv
import html
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from .models import STATUS_NAMES, DictRow

INVISIBLES = re.compile("[\u200b-\u200f\u202a-\u202e\u2060-\u2064\u2066-\u2069\ufeff]")
WHITESPACE = re.compile(r"\s+")
PUNCT_FOLD = str.maketrans(
    {
        "‘": "'",
        "’": "'",
        "‚": "'",
        "‛": "'",
        "“": '"',
        "”": '"',
        "„": '"',
        "–": "-",
        "—": "-",
        "−": "-",
        "\u00a0": " ",
        "\u2007": " ",
        "\u202f": " ",
        "\u2009": " ",
    }
)
TIER_NAMES = ["exact", "normalized", "entity_folded", "casefolded"]


def unescape_all(value: str) -> str:
    for _ in range(3):
        decoded = html.unescape(value)
        if decoded == value:
            break
        value = decoded
    return value


def norm_tiers(value: str | None) -> list[str]:
    value = value or ""
    exact = value
    normalized = INVISIBLES.sub("", unicodedata.normalize("NFC", value)).strip()
    entity_folded = WHITESPACE.sub(
        " ", unescape_all(normalized).translate(PUNCT_FOLD)
    ).strip()
    return [exact, normalized, entity_folded, entity_folded.casefold()]


def require_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX files") from exc
    return openpyxl


@dataclass(frozen=True)
class SheetRow:
    number: int
    row_id: int | None
    source_text: str
    target_text: str


@dataclass
class Match:
    sheet_row: SheetRow
    target: DictRow | None = None
    tier: int | None = None
    outcome: str = "unmatched"
    note: str = ""


def export_dictionary(
    path: str | Path,
    rows: list[DictRow],
    source_language: str,
    target_language: str,
) -> None:
    headers = [
        "row_id",
        "source_language",
        "target_language",
        "source_text",
        "target_text",
        "status",
        "source_hash",
    ]
    data = [
        [
            row.id,
            source_language,
            target_language,
            row.original,
            row.translated or "",
            STATUS_NAMES.get(row.status, row.status),
            row.source_hash,
        ]
        for row in rows
    ]
    destination = Path(path)
    if destination.suffix.casefold() in (".csv", ".tsv"):
        delimiter = "\t" if destination.suffix.casefold() == ".tsv" else ","
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerow(headers)
            writer.writerows(data)
        return

    openpyxl = require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "translations"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
    for values in data:
        sheet.append(values)
    for column, width in (
        (1, 10),
        (2, 18),
        (3, 18),
        (4, 70),
        (5, 70),
        (6, 20),
        (7, 68),
    ):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for record in sheet.iter_rows(min_row=2):
        record[3].alignment = Alignment(wrap_text=True, vertical="top")
        record[4].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(destination)


def _resolve_column(
    headers: list[str], hint: str | None, aliases: tuple[str, ...]
) -> int | None:
    if hint:
        candidate = hint.strip()
        for index, header in enumerate(headers):
            if header.casefold() == candidate.casefold():
                return index
        if candidate.isdigit():
            return int(candidate) - 1
        if re.fullmatch(r"[A-Za-z]{1,3}", candidate):
            number = 0
            for char in candidate.upper():
                number = number * 26 + ord(char) - 64
            return number - 1
        raise ValueError(f"cannot resolve spreadsheet column {hint!r}")
    for index, header in enumerate(headers):
        normalized = re.sub(r"[^a-z0-9]+", "_", header.casefold()).strip("_")
        if normalized in aliases:
            return index
    return None


def load_sheet(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
) -> list[SheetRow]:
    source_path = Path(path)
    if source_path.suffix.casefold() in (".csv", ".tsv"):
        delimiter = "\t" if source_path.suffix.casefold() == ".tsv" else ","
        with source_path.open(encoding="utf-8-sig", newline="") as handle:
            raw = [list(row) for row in csv.reader(handle, delimiter=delimiter)]
    else:
        openpyxl = require_openpyxl()
        workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"worksheet {sheet_name!r} not found; available: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name or workbook.sheetnames[0]]
        raw = [list(row) for row in worksheet.iter_rows(values_only=True)]
        workbook.close()
    raw = [row for row in raw if any(cell not in (None, "") for cell in row)]
    if len(raw) < 2:
        raise ValueError("spreadsheet has no data rows")
    headers = [str(cell or "").strip() for cell in raw[0]]
    source_index = _resolve_column(
        headers,
        source_column,
        ("source", "source_text", "original", "english", "english_original"),
    )
    target_index = _resolve_column(
        headers,
        target_column,
        (
            "target",
            "target_text",
            "translated",
            "translation",
            "german",
            "german_translation",
        ),
    )
    id_index = _resolve_column(headers, None, ("id", "row_id", "dictionary_id"))
    if source_index is None or target_index is None or source_index == target_index:
        raise ValueError(
            "could not identify source and target columns; pass --source-col and --target-col"
        )

    output: list[SheetRow] = []
    for number, row in enumerate(raw[1:], start=2):
        raw_id = (
            str(row[id_index])
            if id_index is not None
            and id_index < len(row)
            and row[id_index] is not None
            else ""
        ).strip()
        row_id = (
            int(float(raw_id))
            if raw_id and re.fullmatch(r"\d+(?:\.0+)?", raw_id)
            else None
        )
        source_text = (
            str(row[source_index])
            if source_index < len(row) and row[source_index] is not None
            else ""
        )
        target_text = (
            str(row[target_index])
            if target_index < len(row) and row[target_index] is not None
            else ""
        )
        if source_text or target_text:
            output.append(SheetRow(number, row_id, source_text, target_text))
    return output


def match_sheet(
    sheet_rows: list[SheetRow],
    dictionary: list[DictRow],
    *,
    max_tier: int = 2,
    skip_translated: bool = True,
) -> list[Match]:
    by_id = {row.id: row for row in dictionary}
    indexes: list[dict[str, list[DictRow]]] = [defaultdict(list) for _ in TIER_NAMES]
    for row in dictionary:
        for tier, key in enumerate(norm_tiers(row.original)):
            indexes[tier][key].append(row)
    output: list[Match] = []
    claimed: dict[int, Match] = {}
    for sheet_row in sheet_rows:
        match = Match(sheet_row)
        if sheet_row.row_id is not None:
            candidate = by_id.get(sheet_row.row_id)
            if candidate and candidate.original == sheet_row.source_text:
                match.target = candidate
                match.tier = 0
            elif candidate:
                match.outcome = "stale_source"
                match.note = "row ID exists but exact source text changed"
                output.append(match)
                continue
        if match.target is None:
            keys = norm_tiers(sheet_row.source_text)
            for tier in range(min(max_tier, len(TIER_NAMES) - 1) + 1):
                candidates = {row.id: row for row in indexes[tier].get(keys[tier], [])}
                if not candidates:
                    continue
                empty = [row for row in candidates.values() if not row.has_translation]
                match.target = (empty or list(candidates.values()))[0]
                match.tier = tier
                if len(candidates) > 1:
                    match.note = f"{len(candidates)} rows share this source; selected id {match.target.id}"
                break
        if match.target is None:
            output.append(match)
            continue
        target_text = sheet_row.target_text
        if not target_text.strip():
            match.outcome = "skipped"
            match.note = "target cell is empty"
        elif skip_translated and match.target.has_translation:
            match.outcome = "already_translated"
            match.note = "existing translation preserved"
        elif (match.target.translated or "") == target_text:
            match.outcome = "unchanged"
        elif match.target.id in claimed:
            match.outcome = "conflict"
            match.note = f"another spreadsheet row already targets id {match.target.id}"
        else:
            match.outcome = "translated"
            claimed[match.target.id] = match
        output.append(match)
    return output
